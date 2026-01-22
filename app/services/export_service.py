import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_
from app.models.group import LedgerRecord, GroupConfig
from datetime import date, datetime, time, timedelta
import io

async def generate_group_ledger(session: AsyncSession, group_id: int, query_date: date) -> io.BytesIO:
    # 1. Fetch Data (Using LedgerRecord)
    # Logic: 4AM to 4AM next day
    start_time = datetime.combine(query_date, time(4, 0))
    end_time = start_time + timedelta(hours=24)
    
    stmt = select(LedgerRecord).where(
        and_(
            LedgerRecord.group_id == group_id,
            LedgerRecord.created_at >= start_time,
            LedgerRecord.created_at < end_time
        )
    ).order_by(LedgerRecord.created_at)
    
    result = await session.execute(stmt)
    transactions = result.scalars().all()
    
    # Separate types
    deposits = [t for t in transactions if t.type == "deposit"]
    payouts = [t for t in transactions if t.type == "payout"]
    
    # 2. Create Workbook
    wb = openpyxl.Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    def style_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

    # Sheet 1: 入款 (Deposits)
    ws1 = wb.active
    ws1.title = "入款明细"
    style_header(ws1, ["时间", "金额", "操作人", "备注/原始指令"])
    
    total_deposit = 0.0
    for t in deposits:
        ws1.append([t.created_at.strftime('%H:%M:%S'), t.amount, t.operator_name, t.original_text])
        total_deposit += t.amount
    
    # Sheet 2: 下发 (Payouts)
    ws2 = wb.create_sheet("下发明细")
    style_header(ws2, ["时间", "金额", "操作人", "备注/原始指令"])
    
    total_payout = 0.0
    for t in payouts:
        ws2.append([t.created_at.strftime('%H:%M:%S'), t.amount, t.operator_name, t.original_text])
        total_payout += t.amount

    # Sheet 3: 汇总 (Summary)
    ws3 = wb.create_sheet("每日汇总")
    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15
    
    # Get Config for Fee
    stmt_config = select(GroupConfig).where(GroupConfig.group_id == group_id)
    res_config = await session.execute(stmt_config)
    config = res_config.scalars().first()
    fee_percent = config.fee_percent if config else 0.0
    
    # Calculate Fees
    fee = total_deposit * (fee_percent / 100.0)
    should_pay = total_deposit - fee
    pending_pay = should_pay - total_payout
    
    summary_data = [
        ("日期", str(query_date)),
        ("入款总额", total_deposit),
        ("费率", f"{fee_percent}%"),
        ("手续费", fee),
        ("应下发金额", should_pay),
        ("实际下发", total_payout),
        ("未下发/结余", pending_pay)
    ]
    
    for row in summary_data:
        ws3.append(row)
        
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
